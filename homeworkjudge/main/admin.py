from django.contrib import admin
from .models import Course,Student,Course_Student,Homework,Problem,Result
import openpyxl
from datetime import datetime
from django.http import HttpResponse
# Register your models here.
class Course_admin(admin.ModelAdmin):
    list_display = ('courseId','courseName')
    def save_model(self, request, obj, form, change):
        xml = obj.studentList
        wb  = openpyxl.load_workbook(xml)
        sheet = wb.worksheets[0]
        print(sheet)
        newAddStudentCount = 0
        newAddStudentList = []
        obj.save()
        for i in list(sheet.rows):
            id,name = i[0].value,i[1].value
            try:
                studentTempObj = Student.objects.get(studentId=id)
            except: #計算有被加入student表的人數
                studentTempObj = Student.objects.create(studentId=id,studentName=name)
                newAddStudentCount +=1
                newAddStudentList.append([id,name])
            courseStudentTempObj = Course_Student.objects.get_or_create(student=studentTempObj,course=obj)
        print("新增人數" ,newAddStudentCount)
        print("帳號,姓名",newAddStudentList)
            
     
        super().save_model(request, obj, form, change)

class Course_Student_admin(admin.ModelAdmin):
    list_filter = ('course',)
    list_display = ('student','course')
    
class Homework_admin(admin.ModelAdmin):
    list_display = ('homeworkId','course','title','startDateTime','endDateTime')
    def export_xlsx(modeladmin,request,queryset):
        print(queryset)
        for homeworkObj in queryset:
            lst = []
            students = Course_Student.objects.filter(course=homeworkObj.course)
            problems = Problem.objects.filter(homework_id=homeworkObj.homeworkId)
            # print(students)
            # print(problems)
            for i in range(len(students)):
                lst.append([])
                for j in range(len(problems)):
                    result = Result.objects.filter(student_id=students[i].student_id).filter(problem_id=problems[j].problemId).first()            
                    if result:
                        lst[i].append(1 if result.result else 0)
                    else:
                        lst[i].append(0)
                    print(lst[i][j], end = ' ')
                    # print(result)
            import openpyxl
            # from openpyxl.cell import get_column_letter
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            fileName = datetime.now().strftime('%Y%m%d%H%M%S')+'.xlsx'
            response['Content-Disposition'] = 'attachment; filename='+fileName
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = homeworkObj.title

            for i in range(len(problems)):
                c = ws.cell(row=1, column=i+3)
                c.value = problems[i].problemName

            # row_num = 0
            for row in range(len(students)):
                studentObj = students[row].student
                c = ws.cell(row=row+2, column=1)
                c.value = studentObj.studentId
                c = ws.cell(row=row+2, column=2)
                c.value = studentObj.studentName
                for column in range(len(problems)):
                    c = ws.cell(row=row+2, column=column+3)
                    c.value = lst[row][column]

            wb.save(response)
            print('EXCEL 匯出')
            return response
    export_xlsx.short_description = '匯出EXCEL檔'
    actions = [export_xlsx,]

class Problem_admin(admin.ModelAdmin):
    list_display = ('problemId','homework','problemName')

class Result_admin(admin.ModelAdmin):
    list_filter = ('student',)

class Student_admin(admin.ModelAdmin):
    list_display = ('studentName','studentId')


admin.site.register(Course,Course_admin)
admin.site.register(Student,Student_admin)
admin.site.register(Course_Student,Course_Student_admin)
admin.site.register(Homework,Homework_admin)
admin.site.register(Problem,Problem_admin)
admin.site.register(Result,Result_admin)